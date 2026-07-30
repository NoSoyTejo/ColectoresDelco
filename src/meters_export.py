"""Exportación de lecturas de medidores a Excel (.xlsx)."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import List, Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from protocol import MeterReading


HEADERS = (
    "Índice",
    "Medidor",
    "T1",
    "T2",
    "T3",
    "T4",
    "Total",
    "Fecha lectura",
    "Display",
    "Estado",
)


def _num(val: Optional[float]):
    return None if val is None else round(float(val), 2)


def _display_label(on: Optional[bool]) -> str:
    if on is True:
        return "ON"
    if on is False:
        return "OFF"
    return ""


def export_meters_xlsx(
    path: str | Path,
    readings: Sequence[MeterReading],
    *,
    source: str = "",
    collector_count: Optional[str] = None,
    notes: str = "",
) -> Path:
    """
    Genera un .xlsx con hoja Lecturas (tabla) y hoja Resumen.
    """
    out = Path(path)
    if out.suffix.lower() != ".xlsx":
        out = out.with_suffix(".xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Lecturas"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )
    alt_fill = PatternFill("solid", fgColor="E8F0F8")
    num_align = Alignment(horizontal="right", vertical="center")
    text_align = Alignment(horizontal="left", vertical="center")

    for col, title in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin

    for idx, reading in enumerate(readings, start=1):
        row = idx + 1
        values = (
            idx,
            reading.meter_id.lstrip("0") or reading.meter_id,
            _num(reading.t1),
            _num(reading.t2),
            _num(reading.t3),
            _num(reading.t4),
            _num(reading.display_total),
            reading.datetime_text or "",
            _display_label(reading.display_on),
            reading.status or "",
        )
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin
            if col in (3, 4, 5, 6, 7) and isinstance(value, (int, float)):
                cell.number_format = "#,##0.00"
                cell.alignment = num_align
            elif col == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = text_align
            if idx % 2 == 0:
                cell.fill = alt_fill

    n = len(readings)
    if n > 0:
        table = Table(displayName="LecturasMedidores", ref=f"A1:J{n + 1}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    widths = (8, 16, 12, 12, 12, 12, 12, 20, 10, 10)
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22

    # --- Resumen ---
    summary = wb.create_sheet("Resumen", 0)
    title_font = Font(bold=True, size=14, color="1F4E79")
    label_font = Font(bold=True)
    summary["A1"] = "Proyecto Colectores — Exportación de medidores"
    summary["A1"].font = title_font
    summary.merge_cells("A1:B1")

    exported_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    rows_info = [
        ("Fecha exportación", exported_at),
        ("Origen / conexión", source or "—"),
        ("Cantidad reportada (O)", collector_count or "—"),
        ("Filas exportadas", n),
        ("Suma Total (columna)", round(sum((r.display_total or 0.0) for r in readings), 2)),
        ("Notas", notes or "Total = F19 del colector; si F19=0 se usa suma T1–T4 (mono-tarifa)."),
    ]
    for i, (label, value) in enumerate(rows_info, start=3):
        summary.cell(row=i, column=1, value=label).font = label_font
        summary.cell(row=i, column=2, value=value)

    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 70

    wb.save(out)
    return out
